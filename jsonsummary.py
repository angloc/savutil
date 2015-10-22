import exceptions
import os

import openpyxl

class EnhancementWorkbookError (exceptions.Exception):
	pass
	
class EnhancementWorkbook (object):
	def __init__ (self, workbookFilename, mode="readonly"):
		self.workbookFilename = workbookFilename
		self.mode = mode
		if mode == "new":
			if os.path.exists (workbookFilename):
				raise EnhancementWorkbookError,\
					"Can't open workbook: '%s' already exists" %\
						workbookFilename
			self.wb = openpyxl.Workbook ()
		else:
			self.wb = openpyxl.load_workbook (workbookFilename)
			
	def save (self):
		if self.mode == "readonly":
			raise EnhancementWorkbookError,\
				"Can't save read-only workbook: '%s'" %\
					self.workbookFilename
		elif self.mode == "new":
			if os.path.exists (self.workbookFilename):
				raise EnhancementWorkbookError,\
					"Can't save new workbook: '%s' already exists" %\
						self.workbookFilename
		self.wb.save (self.workbookFilename)
			
if __name__ == "__main__":

	import getopt
	import json
	import sys
	import traceback

	optlist, args = getopt.getopt (sys.argv[1:], '')

	for (option, value) in optlist:
		pass

	if len (args) < 1:
		print "--Usage: <inputFile> [<outputFile>]"
		sys.exit (0)

	(root, jsonExt) = os.path.splitext (args [0])
	if not jsonExt: jsonExt = ".json"
		
	# Get JSON information
	try:
		jsonFile = open (root + jsonExt)
		jsonData = json.loads (jsonFile.read ())
		jsonFile.close ()
	except exceptions.Exception, e:
		print "--Can't load JSON file (%s)" % e
		traceback.print_exc ()
		sys.exit (0)
		
	totalCount = jsonData ["total_count"]
		
	# Open workbook
	
	if len (args) > 1:
		workbookFilename = args [1]
		workRoot, workExt = os.path.splitext (args [1])
		if not workExt: workExt = ".xlsx"
	else:
		workRoot = root
		workExt = ".xlsx"
	try:
		ewb = EnhancementWorkbook (workRoot + workExt, "new")
	except exceptions.Exception, e:
		print "--Error: %s" % e
		sys.exit (0)
		
	wb = ewb.wb
	distributions = wb.active
	distributions.title = "Input Distributions"
	wb.create_sheet (0, title="Input Lists")
	lists = wb.active
	wb.create_sheet (0, title="Input Map")
	map = wb.active

	map.append ((
		'Name',
		'Type',
		'List',
		'Frequency',
		'Present',
		'Missing',
		'Values',
		'Title'))
	variableSequence = jsonData ["variable_sequence"]
	for variableName in variableSequence:
		variable = jsonData ["variables"] [variableName]
		distribution = variable ["distribution"]
		map.append ((
			variable ["name"],
			variable ["json_type"],
			variable.get ("code_list_name"),
			distribution ["frequency_type"],
			distribution ["non_missing_frequency"],
			distribution ["missing_frequency"],
			distribution ["unique_values"],
			variable.get ("title")
		))
		
	lists.append (("Name", "Code", "Label"))
	codeLists = jsonData.get ("code_lists")
	if codeLists:
		for codeListName in sorted (codeLists.keys ()):
			codeList = codeLists [codeListName]
			sequence = codeList ["sequence"]
			table = codeList ["table"]
			allInteger = True
			for code in sequence:
				if not code.isdigit ():
					allInteger = False
					break
			for code in sequence:
				if allInteger:
					outputCode = int (code)
				else:
					outputCode = code
				lists.append ((codeListName, outputCode, table [code]))

	distributions.append (("Name", "Frequency", "Measure", "Value", "Label"))
	distributions.append (("", totalCount, "All records"))
	variableSequence = jsonData ["variable_sequence"]
	emptyList = {}
	for variableName in variableSequence:
		listName = variable.get ("code_list_name")
		if listName:
			listObject = jsonData ["code_lists"].get (listName)
			listTable = listObject ["table"]
		else:
			listTable = emptyList
		variable = jsonData ["variables"] [variableName]
		jsonType = variable ["json_type"]
		distribution = variable ["distribution"]
		distributions.append ((
			variableName,
			"",
			"Title",
			variable.get ("title")
		))
		distributions.append ((
			variableName,
			distribution ["non_missing_frequency"],
			"Records"
		))
		uniqueValue = None
		if distribution ["frequency_type"] == "constant":
			uniqueValue = distribution ["min_value"]
		distributions.append ((
			variableName,
			distribution ["frequency_type"],
			"Distribution",
			uniqueValue
		))
		if distribution ["frequency_type"] == "empty":
			continue
		if distribution ["missing_frequency"] != 0:
			distributions.append ((
				variableName,
				distribution ["missing_frequency"],
				"Missing"
			))
		if distribution ["frequency_type"] not in ("constant", "id", "unique"):
			distributions.append ((
				variableName,
				distribution ["unique_values"],
				"Unique values"
			))
			if distribution ["unique_frequency"]:
				distributions.append ((
					variableName,
					distribution ["unique_frequency"],
					"Unique frequency"
				))
			distributions.append ((
				variableName,
				distribution ["modal_frequency"],
				"Mode",
				distribution ["modal_value"],
				listTable.get (distribution ["modal_value"])
			))
		if distribution ["frequency_type"] != "constant":
			distributions.append ((
				variableName,
				None,
				"Minimum",
				distribution ["min_value"],
				listTable.get (distribution ["min_value"])
			))
			distributions.append ((
				variableName,
				None,
				"Maximum",
				distribution ["max_value"],
				listTable.get (distribution ["max_value"])
			))
		
		if distribution ["frequency_type"] == "variable":
			distributionTable = distribution ["distribution"]
			keys = distributionTable.keys ()
			if jsonType == "decimal":
				keys.sort (lambda x, y: cmp (float (x), float (y)))
			elif jsonType == "integer":
				keys.sort (lambda x, y: cmp (int (x), int (y)))
			else:
				keys.sort ()
			listedCount = 0
			for key in keys:
				count = distributionTable [key]
				if jsonType == "decimal":
					keyValue = float (key) 
				elif jsonType == "integer":
					keyValue = int (key)
				else:
					keyValue = key
				if count > 1:
					distributions.append ((
						variableName,
						count,
						None,
						keyValue,
						listTable.get (key)
					))
				listedCount += count
			otherCount = distribution ["non_missing_frequency"] - listedCount
			if otherCount:
				distributions.append ((
					variableName,
					otherCount,
					"Others"
				))

	# Save workbook
	
	ewb.save ()
